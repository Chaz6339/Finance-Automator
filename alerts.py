import tkinter as tk
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class Alert():

    alertList = []

    def __init__(self, alertNumber, isHourNumber, rowNum, columnNum, currentSheet, promptVars):
        self.currentSheet = currentSheet
        self.alertNumber = alertNumber
        self.isHourNum = isHourNumber
        self.rowNum = rowNum
        self.columnNum = columnNum
        self.promptVars = promptVars
        self.message = self.gatherAlertInfo()
        if not any(alert.message == self.message for alert in Alert.alertList): #Handles duplicate values coming up for the time entry cell and the hour value cell
            Alert.alertList.append(self)
        else:
            print("Alert there already")
        print("all goood!")



    def findDay(self):
        cellVal = self.currentSheet.cell(row = 2, column = self.columnNum).value

        match cellVal:
            case "Mon":
                return "Monday"
            case "Tue":
                return "Tuesday"
            case "Wed":
                return "Wednesday"
            case "Thu":
                return "Thursday"

    def findPeriod(self):
        return self.currentSheet.cell(row = self.rowNum, column=1).value


    def gatherAlertInfo(self):
        dayWeek = self.findDay()
        period = self.findPeriod()

        alertMessage = "Alert " + str(self.alertNumber) + ":" + "\nYou seem to have missed an entry for " + dayWeek +  "\n during the week of " + period + ". \nWould you like to make an entry?"
        return alertMessage
        #print(dayWeek, period)
    

    @classmethod
    def getAllAlerts(cls):
        return cls.alertList

