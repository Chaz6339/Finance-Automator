from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import tkinter as tk
import re
from datetime import datetime
from alerts import Alert

class helperFunctions:
    def __init__(self, promptVars):
        self.promptVars = promptVars
        self.path = "/Users/chaz/Desktop/Chuck/Test_Finances.xlsx"
        self.currentSheet = 'Job2'
        self.currentWorkBook = self.openFile()
        self.alertNumber = self.checkForNull()
        print("here")




    def openFile(self):
        try:
            file = load_workbook(self.path)
        except FileNotFoundError:
            errorRoot = tk.Tk()
            errorRoot.title("No file found")
            errorRoot.geometry("200x100")
            tk.Label(errorRoot, text = "No file found!").pack()

#-----------------------------Time-Related-Helpers---------------------------------------#
    @staticmethod 
    def totalTime(totalHours, totalMin):
        totalTime = round(totalHours + (totalMin/60), 2)
        return totalTime
    print("TESTING TOTALTIME: " + str(totalTime(2, 15)))
        
    @staticmethod
    def realTotalMinHours(startHour, endHour, startMin, endMin):
        if (endHour < startHour or (endHour == startHour and endMin < startMin)):
            endHour += 12
        totHours = endHour-startHour
        totalMin = endMin - startMin

        if totalMin < 0:
            totHours -= 1
            totalMin += 60

        # if (endMin-startMin < 0):
        #     totHours = endHour-startHour - 1
        #     totalMin = 60 + (endMin - startMin)
        # else:
        #     totalMin = endMin - startMin

        return totalMin, totHours
        
    
#-------------------------Getting-Row-And-Column-Helpers---------------------------------#

    def getRowDate(self, sheet, dayObj, selectedDay):
        """Iterates through the first column (A) in order to find the row (daterange)"""
        col_index = column_index_from_string('A')
        pattern = re.compile(r"(\d{1,2}\/\d{1,2})")
        for col_cells in sheet.iter_cols(min_col=col_index, max_col=col_index):
                for cell in col_cells:
                    if cell.value and re.search(pattern, str(cell.value).strip()):
                         #rowNum
                        #print("CELL VAL DATE:", cell.value)
                        if self.dateInRange(cell.value, sheet, dayObj):#Parses the dates and checks whether the current date is in date range
                            rowNum = helperFunctions.splitterRow(cell) #rowNum from splitter
                            print("yessir" + str(rowNum))
                            return int(rowNum)
                        else:
                            print("Date was not in any range")

        

    @staticmethod
    def getColDay(currentSheet, dayToday):
        """Iterates Through the row (days of the week) and returns the proper y-cord (column #)"""
        pattern = re.compile("^[a-zA-Z]{3}$")
        for rowDay in currentSheet[2]:
            print(rowDay)
            if rowDay.value and re.search(pattern, str(rowDay.value).strip()):
                #print("a" + str(rowDay.value) + todayDay)
                if (str(rowDay.value) in dayToday):
                    colNum = helperFunctions.splitterCol(rowDay)
                    print("COL NUM IS" + str(colNum))
                    return int(colNum) #colNUm
                else:
                    print("Not right row YET")


    @staticmethod
    def dateInRange(cellValueDate, currentSheet, theDateObj):
        """Called from getRow(1).
        validates which row the dated hours go to"""
        year = "25"
        startDateString, endDateString = cellValueDate.split("-")
        startDate = datetime.strptime(f"{startDateString}/{year}", "%m/%d/%y")
        endDate = datetime.strptime(f"{endDateString}/{year}", "%m/%d/%y")
        if startDate <= theDateObj <= endDate: #needs todayDateObj passed into middle value
            print(startDate, endDate)
            return True #returns a bool value
        else:
            return False
        
#----------------------------Searching for empty Vals----------------------#
    def checkForNull(self):
        alertNumber = 0
        print("working")
        file_path = "/Users/chaz/Desktop/Chuck/Test_Finances.xlsx"
        file = load_workbook(file_path)
        currentSheet = file['Job2']
        for row in currentSheet.iter_rows(min_row=3, max_row=32, min_col=2, max_col=5): #max_col ---> time ranges (##:##-##:##) [Like the 'max length of the row'] | max_row ---> 'max length of the column
            #print(row)
            for cell in row:
                print(cell.row)
                if (cell.row % 2 == 0):
                    isHourNumber = True
                    continue   
                isHourNumber = False
                if (cell.value == None):
                    
                    print(cell.row, cell.column)
                    alertNumber+= 1
                    Alert(alertNumber, isHourNumber, cell.row, cell.column, currentSheet, self.promptVars)
        return alertNumber
                    
        

#----------------------------Other-Helpers-----------------------------------#


    def splitterRow(rawCol):
        everythingElse, colRaw = str(rawCol).split(".")
        rowNumber = colRaw[1:-1]
        
        print("row col is " + rowNumber)
        #print("splitterColumn is " + str(colNumber))
        try:
            return rowNumber
        except ValueError:
            print(f"Invalid row number extracted: '{rowNumber}")
            return None
    
    def splitterCol(rowDay):
        everythingElse, rowRaw = str(rowDay).split(".")
        rowNumberAsLet = rowRaw[0]
        colNum = (ord(rowNumberAsLet)-64)
        print("row raw is " + rowRaw)
        print("splitter is " + str(colNum))
        return colNum
    

    