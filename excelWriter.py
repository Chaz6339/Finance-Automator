
import re
from openpyxl import load_workbook
import tkinter as tk
from openpyxl.utils import column_index_from_string
import datetime as dt
from datetime import datetime
from helpers import helperFunctions
import math
import os
import sys


class WritingToExcel:
  

    def __init__(self, promptVars, hourStart, minuteStart, hourEnd, minuteEnd, dayObj, selectedDay, logRoot, labelFrame):
        self.startHour = int(hourStart)
        self.startMin = int(minuteStart)
        self.endHour = int(hourEnd)
        self.endMin = int(minuteEnd)
        self.dayObj = dayObj
        self.selectedDay = selectedDay
        self.logRoot = logRoot
        self.labelFrame = labelFrame
        self.promptVar = promptVars
        self.helper = helperFunctions(self)
        self.totalMin, self.totalHours = self.helper.realTotalMinHours(self.startHour, self.endHour, self.startMin, self.endMin)#use self as a parameter?
        self.totalTime = self.helper.totalTime(self.totalHours, self.totalMin)
        print("here")
    

    @staticmethod
    def checkSuccess(cellTimesString, totalTime, currentSheet, file, file_path, rowNum, colNum, mainFrame, smallerFrame, promptVars, enterTimeFrame, fromAlert = False):

        """This method 'checks' that the cell the user just input for is actually changed, and then saved.
            the 'closeWindow' and 'logAgain' functions are within checkSuccess as they are only needed/accessed here. """


        def closeWindow(seconds):
            """gives a 10 second countdown for window to close"""
            if seconds <= 10:
                closingMessage = f"The window will close in {10 - seconds} seconds" #changed from 5 seconds to 10 seconds to allow for more time to read 'stats'
                closingSuccess.config(text = closingMessage)
                promptVars.app.root.after(1000, lambda: closeWindow(seconds + 1))
            else:
                    promptVars.app.root.destroy()

        def logAgain():
            #for widget in mainFrame.winfo_children():
                #widget.destroy()
            #for widget in smallerFrame.winfo_children():
                 #widget.destroy()
            promptVars.resetUserInterface()

 #------------------------GUI--------------------------#
        if fromAlert:
            for widget in mainFrame.winfo_children():
                if hasattr(widget, 'my_tag') and widget.my_tag == "theq":
                    widget.destroy()



        successMessage = "Success! You logged the time(s): " + str(cellTimesString)
        secondHoursMes = "The total hours logged for the day are: " + str(totalTime) + "!"
        seconds = 0
        closingMessage = "The window will close in " + str(seconds) + "seconds"
        print("checking with vals row,col ",rowNum, colNum)
        if currentSheet.cell(row = rowNum, column = colNum).value == cellTimesString:
            for widget in smallerFrame.winfo_children():
                widget.destroy()
                print("smaller frame widgets destruction attempt")
                #####
            for widget in mainFrame.winfo_children():
                    if hasattr(widget, 'my_tag'): #and widget.my_tag == "theq":
                        widget.destroy()
                    print("Attempt was made to destroy the question")
            if enterTimeFrame != None:
                for widget in enterTimeFrame.winfo_children():
                    widget.destroy()
                    print("print time frame widgets destruct attempt")
                ###adding this bit
                for widget in mainFrame.winfo_children():
                    widget.destroy()
                    print("Attempt was made to destroy the question")
            #for widget in mainFrame.winfo_children(): ---->>> Was causing smallerFrame to be destroyed somehow?
                    #widget.destroy()
            successHeader = tk.Label(smallerFrame, text = successMessage
                                        , font = ("Courier", 13), bg='white')
            secondHoursSuccess = tk.Label(smallerFrame, text = secondHoursMes, font = ("Courier", 13), bg='white')
            againButton = tk.Button(smallerFrame, text = "Log more time", highlightbackground='white', command=logAgain) #turn to self
            closingSuccess = tk.Label(smallerFrame, text= closingMessage, font = ("Comic Sans MS", 13), bg='white')
            secondHoursSuccess.pack()
            successHeader.pack()
            againButton.pack()
            closingSuccess.pack()
            closeWindow(seconds)         
        else:
            for widget in smallerFrame.winfo_children():
                widget.destroy()
            NosuccessHeader = tk.Label(smallerFrame, text = "Something went wrong, no hours gained, no hours lost!")
            NosuccessHeader.pack()

        print("SUCCESS")
        file.save(file_path)

        
    @staticmethod
    def writeToCell(startHour, endHour, startMin, endMin, promptVar,
                        mainFrame, currentFrame, filePath = "/Users/chaz/Desktop/Chuck/Test_Finances.xlsx",
                        sheetName = 'Job2', fromInstance = False, helper = None,
                         selectedDay = None, dayObject = None, rowNum = None, colNum = None, enterTimeFrame = None, fromAlert = False):
        try:
            hourS = int(startHour)
            hourE = int(endHour)
            minS = int(startMin)
            minE = int(endMin)
        except:
             hourS = hourE = minS = minE = 0

        
        totalMin, totalHours = helperFunctions.realTotalMinHours(hourS, hourE, minS, minE)
        totalTime = helperFunctions.totalTime(totalHours, totalMin)
        cellTimesString = f"{hourS}:{minS:02}-{hourE}:{minE:02}" #if totalTime != 0 else 0

        
        try:
            file = load_workbook(filePath)
            currentSheet = file[sheetName]
            
            if fromInstance and helper is not None and selectedDay and dayObject: #checks to see whether writing in response to an alert or not
                colNumber = helper.getColDay(currentSheet, selectedDay)
                rowNumber = helper.getRowDate(currentSheet, dayObject, selectedDay)
                print("not using alerted vals")
                print(f"rowNumber: {rowNumber}, colNumber: {colNumber}")
            else:
                 colNumber = colNum
                 rowNumber = rowNum
                 print("using alerted row and col")
                 print(f"rowNumber: {rowNumber}, colNumber: {colNumber}")
            print(f"rowNumber: {rowNumber}, colNumber: {colNumber}")
            currentSheet.cell(row = rowNumber, column = colNumber, value = cellTimesString) #cellTimesString)
            currentSheet.cell(row = rowNumber + 1, column = colNumber, value = float(round(totalTime, 2))) #int(totalTime)
            WritingToExcel.checkSuccess(cellTimesString, totalTime, currentSheet, file, filePath, rowNumber, colNumber, mainFrame, currentFrame, promptVar, enterTimeFrame, fromAlert)
        except KeyError:
             pass
        # except Exception as e:
        #      something = type(rowNum)
        #      print (f"Something went wrong: {e} {something}")


            
         
        
